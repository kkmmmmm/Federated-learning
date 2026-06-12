"""Write results to Excel, reproducing the published figure-data layouts.

Produces:
* ``AUROC.xlsx`` / ``calibration_slope.xlsx`` / ``calibration_intercept.xlsx``
  -- the per-validation-region "block" layout used by the published Figures
  1-3 and S1-S9 (one sheet per penalty), so the corrected numbers drop straight
  into the existing chart templates.
* ``PCA.xlsx`` -- the Figure 4 layout (L1/L2/EN side by side, no-reg below).
* ``results.xlsx`` -- a comprehensive workbook: Table 2 / S2 (within-region
  mean(SD)), raw within-region folds, between-region long format, FL
  convergence and model coefficients.
"""
from __future__ import annotations

import os
import numpy as np
import pandas as pd

from . import config as C

PEN_SUFFIX = {"none": "noreg", "l1": "L1", "l2": "L2", "elasticnet": "EN"}
METRIC_PREFIX = {"AUROC": "AUROC", "CalibrationSlope": "slope",
                 "CalibrationIntercept": "intercept"}


def _source_order(r: int):
    foreign = [i for i in C.REGIONS if i != r]
    return foreign + ["Centralized", "FL"]


def _block_sheet(ws, between_pen: pd.DataFrame, metric: str):
    """Fill one worksheet with the 16 per-region blocks for a single metric."""
    row = 1
    for r in C.REGIONS:
        sub = between_pen[between_pen["ValidationRegion"] == r].copy()
        sub["Source"] = sub["Source"].astype(str)
        sub = sub.set_index("Source")
        order = [str(i) for i in C.REGIONS if i != r] + ["Centralized", "FL"]
        headers = [f"region{s}" if s.isdigit() else s for s in order]
        pts, los, ups, eu, el = [], [], [], [], []
        for s in order:
            p = sub.loc[s, metric]
            lo = sub.loc[s, f"{metric}_lower"]
            up = sub.loc[s, f"{metric}_upper"]
            pts.append(p); los.append(lo); ups.append(up)
            eu.append(up - p); el.append(p - lo)

        ws.cell(row, 1, f"region{r}")
        ws.cell(row + 1, 1, "category")
        ws.cell(row + 2, 1, metric)
        ws.cell(row + 3, 1, "ci_lower")
        ws.cell(row + 4, 1, "ci_upper")
        ws.cell(row + 5, 1, "err_upper")
        ws.cell(row + 6, 1, "err_lower")
        for j, h in enumerate(headers, start=2):
            ws.cell(row + 1, j, h)
            ws.cell(row + 2, j, pts[j - 2])
            ws.cell(row + 3, j, los[j - 2])
            ws.cell(row + 4, j, ups[j - 2])
            ws.cell(row + 5, j, eu[j - 2])
            ws.cell(row + 6, j, el[j - 2])
        row += 8


def write_block_workbook(between: pd.DataFrame, metric: str, path: str):
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    present = [p for p in C.PENALTIES if (between["Penalty"] == p).any()]
    for pen in present:
        ws = wb.create_sheet(f"{METRIC_PREFIX[metric]}_{PEN_SUFFIX[pen]}")
        _block_sheet(ws, between[between["Penalty"] == pen], metric)
    wb.save(path)


def write_pca_workbook(pca: dict[str, pd.DataFrame], path: str):
    """Reproduce the Figure 4 layout: L1/L2/EN side by side, no-reg block below."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "fig4_PCA"

    # top block: L1 (B-D), L2 (G-I), EN (L-N)
    groups = [("l1", 1), ("l2", 6), ("elasticnet", 11)]
    ws.cell(1, 1, "Sheet")
    for pen, c0 in groups:
        if pen not in pca:
            continue
        ws.cell(1, c0 + 1, "Index"); ws.cell(1, c0 + 2, "PC1"); ws.cell(1, c0 + 3, "PC2")
        df = pca[pen]
        for k, (_, rr) in enumerate(df.iterrows()):
            ws.cell(2 + k, c0, {"l1": "L1", "l2": "L2", "elasticnet": "EN"}[pen])
            ws.cell(2 + k, c0 + 1, rr["Index"])
            ws.cell(2 + k, c0 + 2, float(rr["PC1"]))
            ws.cell(2 + k, c0 + 3, float(rr["PC2"]))

    # lower block: no regularization
    base = 22
    if "none" not in pca:
        wb.save(path)
        return
    ws.cell(base, 2, "Index"); ws.cell(base, 3, "PC1"); ws.cell(base, 4, "PC2")
    for k, (_, rr) in enumerate(pca["none"].iterrows()):
        ws.cell(base + 1 + k, 1, "noreg")
        ws.cell(base + 1 + k, 2, rr["Index"])
        ws.cell(base + 1 + k, 3, float(rr["PC1"]))
        ws.cell(base + 1 + k, 4, float(rr["PC2"]))
    wb.save(path)


# --------------------------------------------------------------------------- #
# Comprehensive results workbook
# --------------------------------------------------------------------------- #
def _within_summary(within: pd.DataFrame) -> pd.DataFrame:
    """Mean(SD) AUROC by region x model x penalty (Table 2 / S2 source)."""
    g = (within.groupby(["Penalty", "Region", "Model"])["AUROC"]
         .agg(["mean", "std"]).reset_index())
    g["AUROC_mean_sd"] = g.apply(lambda x: f"{x['mean']:.3f} ({x['std']:.3f})", axis=1)
    return g


def _table_like(within: pd.DataFrame, penalty: str, df_counts) -> pd.DataFrame:
    """Build a Table 2 / S2-style frame for one penalty."""
    s = _within_summary(within)
    s = s[s["Penalty"] == penalty]
    piv = s.pivot(index="Region", columns="Model", values="AUROC_mean_sd")
    piv = piv.reindex(columns=["Local", "Centralized", "FL"])
    piv.insert(0, "n", [df_counts[r]["n"] for r in piv.index])
    piv.insert(1, "Event_%", [round(100 * df_counts[r]["event"], 1) for r in piv.index])
    piv = piv.reset_index().rename(columns={"Region": "Region"})
    return piv


def write_results_workbook(within: pd.DataFrame, between: pd.DataFrame,
                           pca: dict, conv: dict, coeffs: pd.DataFrame,
                           df_counts: dict, path: str):
    present = [p for p in C.PENALTIES if (within["Penalty"] == p).any()]
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        # Table 2 (L1) and Table S2 (all penalties)
        for pen in present:
            name = {"l1": "Table2_L1", "none": "TableS2_noreg",
                    "l2": "TableS2_L2", "elasticnet": "TableS2_EN"}[pen]
            _table_like(within, pen, df_counts).to_excel(xl, name, index=False)
        # raw folds
        within.to_excel(xl, "within_region_folds", index=False)
        # between-region long format
        between.to_excel(xl, "between_region", index=False)
        # PCA
        for pen in present:
            if pen in pca:
                pca[pen].to_excel(xl, f"PCA_{PEN_SUFFIX[pen]}", index=False)
        # FL convergence
        conv_rows = []
        for pen, h in conv.items():
            for k in range(len(h.rounds)):
                conv_rows.append(dict(Penalty=pen, Round=h.rounds[k],
                                      TrainLogLoss=h.train_logloss[k],
                                      TrainAUROC=h.train_auroc[k]))
        pd.DataFrame(conv_rows).to_excel(xl, "FL_convergence", index=False)
        conv_summary = pd.DataFrame([
            dict(Penalty=pen, ConvergedRound=h.converged_round,
                 WallSeconds=round(h.wall_seconds, 3),
                 RecalSeconds=round(getattr(h, "recal_seconds", 0.0), 4),
                 TotalSeconds=round(h.wall_seconds + getattr(h, "recal_seconds", 0.0), 3),
                 RecalDelta=getattr(h, "recal_delta", None),
                 FinalLogLoss=h.train_logloss[-1], FinalAUROC=h.train_auroc[-1],
                 RecalLogLoss=getattr(h, "recal_logloss", None),
                 RecalAUROC=getattr(h, "recal_auroc", None))
            for pen, h in conv.items()])
        conv_summary.to_excel(xl, "FL_convergence_summary", index=False)
        # coefficients
        coeffs.to_excel(xl, "coefficients", index=False)
